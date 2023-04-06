import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from io import BytesIO

class ExcelMarker:
    def get_marked_pricelist(self, uploaded_file, not_full_rows: list, color: str) -> bytes:
        for chunk in uploaded_file.chunks():
            
            wb :openpyxl.Workbook = openpyxl.load_workbook(BytesIO(chunk))
            sheet_names = wb.get_sheet_names()
            sheet:Worksheet = wb.get_sheet_by_name(sheet_names[0])
            for row in not_full_rows:
                for cell in row:
                    sheet[cell.coordinate].fill = PatternFill("solid", fgColor=color)
            uploaded_file.seek(0)
        virtual = BytesIO() 
        wb.save(virtual)

        virtual.seek(0)
        output = virtual.read()
        return output